#!/usr/bin/env python3
"""
Super-Democracy - 班委投票系统
一个简洁高效的班委满意度投票系统，支持多班级管理
"""

import os
import sqlite3
import json
from datetime import datetime, timedelta
from functools import wraps
from io import BytesIO
from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, send_file
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import openpyxl
import openpyxl.styles

# 初始化 Flask 应用
app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'super-democracy-secret-key-2024')
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=7)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# 确保上传目录存在
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# 允许的文件扩展名
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

def allowed_file(filename):
    """检查文件扩展名是否允许"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# 数据库连接函数
def get_db():
    """创建数据库连接"""
    conn = sqlite3.connect('database.db')
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    """初始化数据库"""
    conn = get_db()
    c = conn.cursor()
    
    # 用户表
    c.execute('''CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        class_name TEXT NOT NULL,
        username TEXT UNIQUE NOT NULL,
        password_hash TEXT NOT NULL,
        is_admin INTEGER DEFAULT 0,
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP
    )''')
    
    # 班委岗位表
    c.execute('''CREATE TABLE IF NOT EXISTS positions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        class_name TEXT NOT NULL,
        position_name TEXT NOT NULL,
        member_name TEXT NOT NULL,
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        UNIQUE(class_name, position_name)
    )''')
    
    # 投票记录表
    c.execute('''CREATE TABLE IF NOT EXISTS votes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,
        position_id INTEGER NOT NULL,
        is_satisfied INTEGER NOT NULL,
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (user_id) REFERENCES users(id),
        FOREIGN KEY (position_id) REFERENCES positions(id),
        UNIQUE(user_id, position_id)
    )''')
    
    # 创建索引以提高查询性能
    c.execute('CREATE INDEX IF NOT EXISTS idx_users_class ON users(class_name)')
    c.execute('CREATE INDEX IF NOT EXISTS idx_positions_class ON positions(class_name)')
    c.execute('CREATE INDEX IF NOT EXISTS idx_votes_user ON votes(user_id)')
    
    conn.commit()
    
    # 检查是否有管理员账号，如果没有则创建默认管理员
    c.execute('SELECT COUNT(*) as count FROM users WHERE is_admin = 1')
    if c.fetchone()['count'] == 0:
        admin_hash = generate_password_hash('admin123')
        c.execute('''INSERT INTO users (name, class_name, username, password_hash, is_admin) 
                     VALUES (?, ?, ?, ?, ?)''', 
                  ('系统管理员', '管理员', 'admin', admin_hash, 1))
        conn.commit()
        print("已创建默认管理员账号 - 用户名: admin, 密码: admin123")
    
    conn.close()

# 认证装饰器
def login_required(f):
    """要求登录的装饰器"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('请先登录', 'error')
            return redirect(url_for('login', next=request.url))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    """要求管理员权限的装饰器"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('请先登录', 'error')
            return redirect(url_for('login'))
        
        if not session.get('is_admin'):
            flash('需要管理员权限', 'error')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

# 辅助函数
def get_user_info():
    """获取当前登录用户信息"""
    if 'user_id' in session:
        return {
            'id': session['user_id'],
            'name': session.get('name'),
            'class_name': session.get('class_name'),
            'is_admin': session.get('is_admin', False)
        }
    return None

# 路由定义

@app.route('/')
def index():
    """首页 - 重定向到相应页面"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    if session.get('is_admin'):
        return redirect(url_for('admin_dashboard'))
    else:
        return redirect(url_for('vote'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    """登录页面"""
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        if not username or not password:
            flash('请输入用户名和密码', 'error')
            return render_template('login.html')
        
        conn = get_db()
        c = conn.cursor()
        c.execute('SELECT * FROM users WHERE username = ?', (username,))
        user = c.fetchone()
        conn.close()
        
        if user and check_password_hash(user['password_hash'], password):
            # 设置 session
            session['user_id'] = user['id']
            session['name'] = user['name']
            session['class_name'] = user['class_name']
            session['is_admin'] = bool(user['is_admin'])
            session.permanent = True
            
            flash(f'欢迎回来，{user["name"]}！', 'success')
            
            # 重定向到之前的页面或默认页面
            next_page = request.args.get('next')
            if next_page and next_page.startswith('/'):
                return redirect(next_page)
            
            if user['is_admin']:
                return redirect(url_for('admin_dashboard'))
            else:
                return redirect(url_for('vote'))
        else:
            flash('用户名或密码错误', 'error')
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    """退出登录"""
    session.clear()
    flash('已安全退出', 'success')
    return redirect(url_for('login'))

@app.route('/vote')
@login_required
def vote():
    """投票页面 - 显示本班级的班委"""
    user_info = get_user_info()
    if user_info['is_admin']:
        return redirect(url_for('admin_dashboard'))
    
    conn = get_db()
    c = conn.cursor()
    
    # 获取用户班级的所有班委岗位
    c.execute('''SELECT p.*, 
                 CASE WHEN v.id IS NOT NULL THEN v.is_satisfied ELSE NULL END as user_vote
                 FROM positions p
                 LEFT JOIN votes v ON p.id = v.position_id AND v.user_id = ?
                 WHERE p.class_name = ?
                 ORDER BY p.position_name''', 
              (user_info['id'], user_info['class_name']))
    positions = c.fetchall()
    
    conn.close()
    
    return render_template('index.html', positions=positions, user_info=user_info)

@app.route('/submit_vote', methods=['POST'])
@login_required
def submit_vote():
    """提交投票"""
    user_info = get_user_info()
    if user_info['is_admin']:
        return jsonify({'success': False, 'message': '管理员不能参与投票'})
    
    position_id = request.form.get('position_id')
    is_satisfied = request.form.get('is_satisfied')
    
    if not position_id or is_satisfied is None:
        return jsonify({'success': False, 'message': '参数错误'})
    
    conn = get_db()
    c = conn.cursor()
    
    try:
        # 验证岗位是否属于用户班级
        c.execute('SELECT class_name FROM positions WHERE id = ?', (position_id,))
        position = c.fetchone()
        
        if not position or position['class_name'] != user_info['class_name']:
            return jsonify({'success': False, 'message': '无权对此岗位投票'})
        
        # 插入或更新投票记录
        c.execute('''INSERT OR REPLACE INTO votes (user_id, position_id, is_satisfied) 
                     VALUES (?, ?, ?)''', 
                  (user_info['id'], position_id, int(is_satisfied)))
        conn.commit()
        
        return jsonify({'success': True, 'message': '投票成功！'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'投票失败：{str(e)}'})
    finally:
        conn.close()

# 管理员路由

@app.route('/admin')
@admin_required
def admin_dashboard():
    """管理员仪表板"""
    conn = get_db()
    c = conn.cursor()
    
    # 统计数据
    c.execute('SELECT COUNT(*) as count FROM users WHERE is_admin = 0')
    user_count = c.fetchone()['count']
    
    c.execute('SELECT COUNT(DISTINCT class_name) as count FROM users WHERE is_admin = 0')
    class_count = c.fetchone()['count']
    
    c.execute('SELECT COUNT(*) as count FROM positions')
    position_count = c.fetchone()['count']
    
    c.execute('SELECT COUNT(DISTINCT user_id) as count FROM votes')
    voted_count = c.fetchone()['count']
    
    conn.close()
    
    stats = {
        'user_count': user_count,
        'class_count': class_count,
        'position_count': position_count,
        'voted_count': voted_count,
        'vote_rate': f"{(voted_count / user_count * 100):.1f}" if user_count > 0 else "0"
    }
    
    return render_template('admin/dashboard.html', stats=stats)

@app.route('/admin/users')
@admin_required
def admin_users():
    """用户管理页面"""
    class_filter = request.args.get('class', '')
    
    conn = get_db()
    c = conn.cursor()
    
    # 获取所有班级列表
    c.execute('SELECT DISTINCT class_name FROM users WHERE is_admin = 0 ORDER BY class_name')
    classes = [row['class_name'] for row in c.fetchall()]
    
    # 获取用户列表
    if class_filter:
        c.execute('SELECT * FROM users WHERE is_admin = 0 AND class_name = ? ORDER BY name', 
                  (class_filter,))
    else:
        c.execute('SELECT * FROM users WHERE is_admin = 0 ORDER BY class_name, name')
    
    users = c.fetchall()
    conn.close()
    
    return render_template('admin/users.html', users=users, classes=classes, 
                         current_class=class_filter)

@app.route('/admin/users/add', methods=['POST'])
@admin_required
def admin_add_user():
    """添加用户"""
    name = request.form.get('name')
    class_name = request.form.get('class_name')
    username = request.form.get('username')
    password = request.form.get('password')
    
    if not all([name, class_name, username, password]):
        flash('所有字段都必须填写', 'error')
        return redirect(url_for('admin_users'))
    
    conn = get_db()
    c = conn.cursor()
    
    try:
        password_hash = generate_password_hash(password)
        c.execute('''INSERT INTO users (name, class_name, username, password_hash) 
                     VALUES (?, ?, ?, ?)''', 
                  (name, class_name, username, password_hash))
        conn.commit()
        flash('用户添加成功', 'success')
    except sqlite3.IntegrityError:
        flash('用户名已存在', 'error')
    except Exception as e:
        flash(f'添加失败：{str(e)}', 'error')
    finally:
        conn.close()
    
    return redirect(url_for('admin_users'))

@app.route('/admin/users/<int:user_id>/delete', methods=['POST'])
@admin_required
def admin_delete_user(user_id):
    """删除用户"""
    conn = get_db()
    c = conn.cursor()
    
    try:
        # 删除用户的投票记录
        c.execute('DELETE FROM votes WHERE user_id = ?', (user_id,))
        # 删除用户
        c.execute('DELETE FROM users WHERE id = ? AND is_admin = 0', (user_id,))
        conn.commit()
        
        if c.rowcount > 0:
            flash('用户删除成功', 'success')
        else:
            flash('用户不存在或无法删除', 'error')
    except Exception as e:
        flash(f'删除失败：{str(e)}', 'error')
    finally:
        conn.close()
    
    return redirect(url_for('admin_users'))

@app.route('/admin/users/import', methods=['POST'])
@admin_required
def admin_import_users():
    """批量导入用户"""
    if 'file' not in request.files:
        flash('请选择文件', 'error')
        return redirect(url_for('admin_users'))
    
    file = request.files['file']
    if file.filename == '':
        flash('请选择文件', 'error')
        return redirect(url_for('admin_users'))
    
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        try:
            # 读取 Excel 文件
            workbook = openpyxl.load_workbook(filepath)
            sheet = workbook.active
            
            conn = get_db()
            c = conn.cursor()
            
            success_count = 0
            error_count = 0
            
            # 跳过标题行
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if len(row) >= 4 and all(row[:4]):  # 确保有足够的列且不为空
                    name, class_name, username, password = row[:4]
                    
                    try:
                        password_hash = generate_password_hash(str(password))
                        c.execute('''INSERT INTO users (name, class_name, username, password_hash) 
                                     VALUES (?, ?, ?, ?)''', 
                                  (name, class_name, username, password_hash))
                        success_count += 1
                    except sqlite3.IntegrityError:
                        error_count += 1
                        continue
            
            conn.commit()
            conn.close()
            
            # 删除上传的文件
            os.remove(filepath)
            
            flash(f'导入完成：成功 {success_count} 条，失败 {error_count} 条', 'success')
        except Exception as e:
            flash(f'导入失败：{str(e)}', 'error')
            if os.path.exists(filepath):
                os.remove(filepath)
    else:
        flash('不支持的文件格式', 'error')
    
    return redirect(url_for('admin_users'))

@app.route('/admin/positions')
@admin_required
def admin_positions():
    """岗位管理页面"""
    class_filter = request.args.get('class', '')
    
    conn = get_db()
    c = conn.cursor()
    
    # 获取所有班级列表
    c.execute('SELECT DISTINCT class_name FROM positions ORDER BY class_name')
    classes = [row['class_name'] for row in c.fetchall()]
    
    # 获取岗位列表
    if class_filter:
        c.execute('SELECT * FROM positions WHERE class_name = ? ORDER BY position_name', 
                  (class_filter,))
    else:
        c.execute('SELECT * FROM positions ORDER BY class_name, position_name')
    
    positions = c.fetchall()
    conn.close()
    
    return render_template('admin/positions.html', positions=positions, classes=classes, 
                         current_class=class_filter)

@app.route('/admin/positions/add', methods=['POST'])
@admin_required
def admin_add_position():
    """添加岗位"""
    class_name = request.form.get('class_name')
    position_name = request.form.get('position_name')
    member_name = request.form.get('member_name')
    
    if not all([class_name, position_name, member_name]):
        flash('所有字段都必须填写', 'error')
        return redirect(url_for('admin_positions'))
    
    conn = get_db()
    c = conn.cursor()
    
    try:
        c.execute('''INSERT INTO positions (class_name, position_name, member_name) 
                     VALUES (?, ?, ?)''', 
                  (class_name, position_name, member_name))
        conn.commit()
        flash('岗位添加成功', 'success')
    except sqlite3.IntegrityError:
        flash('该班级的此岗位已存在', 'error')
    except Exception as e:
        flash(f'添加失败：{str(e)}', 'error')
    finally:
        conn.close()
    
    return redirect(url_for('admin_positions'))

@app.route('/admin/positions/<int:position_id>/delete', methods=['POST'])
@admin_required
def admin_delete_position(position_id):
    """删除岗位"""
    conn = get_db()
    c = conn.cursor()
    
    try:
        # 删除相关投票记录
        c.execute('DELETE FROM votes WHERE position_id = ?', (position_id,))
        # 删除岗位
        c.execute('DELETE FROM positions WHERE id = ?', (position_id,))
        conn.commit()
        
        if c.rowcount > 0:
            flash('岗位删除成功', 'success')
        else:
            flash('岗位不存在', 'error')
    except Exception as e:
        flash(f'删除失败：{str(e)}', 'error')
    finally:
        conn.close()
    
    return redirect(url_for('admin_positions'))

@app.route('/admin/statistics')
@admin_required
def admin_statistics():
    """投票统计页面"""
    class_filter = request.args.get('class', '')
    
    conn = get_db()
    c = conn.cursor()
    
    # 获取所有班级列表
    c.execute('SELECT DISTINCT class_name FROM positions ORDER BY class_name')
    classes = [row['class_name'] for row in c.fetchall()]
    
    # 获取统计数据
    if class_filter:
        c.execute('''
            SELECT 
                p.*,
                COUNT(v.id) as total_votes,
                SUM(CASE WHEN v.is_satisfied = 1 THEN 1 ELSE 0 END) as satisfied_votes,
                (SELECT COUNT(*) FROM users WHERE class_name = p.class_name AND is_admin = 0) as class_users
            FROM positions p
            LEFT JOIN votes v ON p.id = v.position_id
            WHERE p.class_name = ?
            GROUP BY p.id
            ORDER BY p.position_name
        ''', (class_filter,))
    else:
        c.execute('''
            SELECT 
                p.*,
                COUNT(v.id) as total_votes,
                SUM(CASE WHEN v.is_satisfied = 1 THEN 1 ELSE 0 END) as satisfied_votes,
                (SELECT COUNT(*) FROM users WHERE class_name = p.class_name AND is_admin = 0) as class_users
            FROM positions p
            LEFT JOIN votes v ON p.id = v.position_id
            GROUP BY p.id
            ORDER BY p.class_name, p.position_name
        ''')
    
    statistics = []
    for row in c.fetchall():
        stat = dict(row)
        if stat['total_votes'] > 0:
            stat['satisfaction_rate'] = (stat['satisfied_votes'] / stat['total_votes']) * 100
        else:
            stat['satisfaction_rate'] = 0
        
        if stat['class_users'] > 0:
            stat['participation_rate'] = (stat['total_votes'] / stat['class_users']) * 100
        else:
            stat['participation_rate'] = 0
        
        statistics.append(stat)
    
    conn.close()
    
    return render_template('admin/statistics.html', statistics=statistics, classes=classes, 
                         current_class=class_filter)

@app.route('/admin/statistics/export')
@admin_required
def export_statistics():
    """导出投票统计报表"""
    class_filter = request.args.get('class', '')
    
    conn = get_db()
    c = conn.cursor()
    
    # 获取统计数据（与统计页面相同的查询）
    if class_filter:
        c.execute('''
            SELECT 
                p.*,
                COUNT(v.id) as total_votes,
                SUM(CASE WHEN v.is_satisfied = 1 THEN 1 ELSE 0 END) as satisfied_votes,
                (SELECT COUNT(*) FROM users WHERE class_name = p.class_name AND is_admin = 0) as class_users
            FROM positions p
            LEFT JOIN votes v ON p.id = v.position_id
            WHERE p.class_name = ?
            GROUP BY p.id
            ORDER BY p.position_name
        ''', (class_filter,))
    else:
        c.execute('''
            SELECT 
                p.*,
                COUNT(v.id) as total_votes,
                SUM(CASE WHEN v.is_satisfied = 1 THEN 1 ELSE 0 END) as satisfied_votes,
                (SELECT COUNT(*) FROM users WHERE class_name = p.class_name AND is_admin = 0) as class_users
            FROM positions p
            LEFT JOIN votes v ON p.id = v.position_id
            GROUP BY p.id
            ORDER BY p.class_name, p.position_name
        ''')
    
    rows = c.fetchall()
    conn.close()
    
    # 创建 Excel 文件
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "投票统计"
    
    # 设置标题
    headers = ['班级', '岗位', '班委姓名', '满意票数', '总票数', '满意度(%)', '班级人数', '参与率(%)']
    ws.append(headers)
    
    # 设置标题行样式
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    
    # 添加数据
    for row in rows:
        total_votes = row['total_votes']
        satisfied_votes = row['satisfied_votes'] or 0
        class_users = row['class_users']
        
        satisfaction_rate = (satisfied_votes / total_votes * 100) if total_votes > 0 else 0
        participation_rate = (total_votes / class_users * 100) if class_users > 0 else 0
        
        ws.append([
            row['class_name'],
            row['position_name'],
            row['member_name'],
            satisfied_votes,
            total_votes,
            f"{satisfaction_rate:.1f}",
            class_users,
            f"{participation_rate:.1f}"
        ])
    
    # 调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # 保存到内存并返回
    from io import BytesIO
    from flask import send_file
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"投票统计_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    if class_filter:
        filename = f"投票统计_{class_filter}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

# 错误处理
@app.errorhandler(404)
def page_not_found(e):
    return render_template('404.html'), 404

@app.errorhandler(500)
def internal_error(e):
    return render_template('500.html'), 500

# 主程序入口
if __name__ == '__main__':
    init_db()
    app.run(host='0.0.0.0', port=5000, debug=True)